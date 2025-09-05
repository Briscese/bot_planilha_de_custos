import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

wb = Workbook()
ws = wb.active
ws.merge_cells('B2:J2')
ws.merge_cells('B3:D7')
ws.merge_cells('E3:F3')
ws.merge_cells('E6:E7')
ws.merge_cells('F6:F7')
ws.merge_cells('H3:J3')
ws.merge_cells('H4:J4')
ws.merge_cells('H5:J5')
ws.merge_cells('H6:J6')
ws.merge_cells('H7:J7')
ws.merge_cells('B8:J8')
ws.merge_cells('B18:J18')
ws.merge_cells

imgito = Image('/img/image.png') 

ws['B2'] = "RELATÓRIO DE DESPESAS - REEMBOLSO"
ws['E3'] = "PERÍODO DAS DESPESAS"
ws['E4'] = "Data Inicial: "
ws['E5'] = "Data Final: "
ws['E6'] = "Dia do Pagamento: "
ws['F6'] = "conforme politica"

ws['G3'] = "NOME: "
ws['G4'] = "CPF OU CNPJ: "
ws['G5'] = "BANCO: "
ws['G6'] = "AGENCIA E/C: "
ws['G7'] = "PIX: "

ws['H3'] = "Rubsney Nascimento"
ws['H4'] = "XXXXXX"
ws['H5'] = "Nubank"

ws['H7'] = "XXXXXX"

ws['B8'] = "Relató de Reembolso"

ws['B9'] = "DATA"
ws['C9'] = "CLIENTE"
ws['D9'] = "TIPO DE DESPESAS"
ws['E9'] = "CHAMADO"
ws['F9'] = "ORIGEM"
ws['G9'] = "DESTINO"
ws['H9'] = "KM RODADO / VALOR"
ws['I9'] = "VALOR"
ws['J9'] = "REF. MAPS"

ws['I17'] = '=SUM(I10:I15)'
ws['B17'] = "TOTAL A RECEBER"

ws['B18'] = "APROVAÇÕES"

ws['C20'] = "NOME COLABORADOR"
ws['C21'] = "ASSINATURA"
ws['C22'] = "DATA"



ws.add_image(imgito, 'B3')

